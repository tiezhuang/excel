# -*- codeing = utf-8 -*-
# @Time : 2020/8/15 13:34
# @Author : 铁壮
# import pandas as pd
from openpyxl import load_workbook

# 获取表格数据赋值给wb
path = r'C:\Users\w\Desktop\66.xlsx'
wb = load_workbook(path)
# 找到工作表
sheet = wb.active

# a = pd.read_csv(r'C:\Users\w\Desktop\test.csv', 'r')
# print(a.head(3))
# 获取单元格 行 列 获取单元格的值
# print(sheet.cell(2, 2).value)
# i 变量 （） 循环开始值 循环结束值
for i in range(2, 22):
    data = sheet.cell(i, 2).value
    print(data)
    # 条件判断
    if isinstance(data, int):

        if data <= 2:
            # 把筛选的数据，新增行写入数据，重复赋值：
            sheet.cell(i, 4).value = "低"
        elif data <= 3:
            sheet.cell(i, 4).value = "中"
        else:
            sheet.cell(i, 4).value = "高"
    else:
        sheet.cell(i, 4).value = "未就业"
# 保存文件
wb.save(path)
